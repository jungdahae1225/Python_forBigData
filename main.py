#!/usr/bin/python3

# HW2)) 20190356 컴퓨터학과 정다해 bigdata homework for excel
'''
update 20211020 for
'''

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename='student.xlsx')
sheet_ranges = wb['Sheet1']
# print(sheet_ranges['A2'].value)

# --update total score--#
row_num = 2
while True:
    mid = sheet_ranges.cell(row=row_num, column=3).value  # 30%
    fin = sheet_ranges.cell(row=row_num, column=4).value  # 35%
    hw = sheet_ranges.cell(row=row_num, column=5).value  # 34%
    att = sheet_ranges.cell(row=row_num, column=6).value  # 1% (just add)
    if mid is None:
        break

    total = mid * 0.3 + fin * 0.35 + hw * 0.34 + att
    row_num += 1

    sheet_ranges.cell(row=row_num - 1, column=7, value=total)

# print(row_num)  # 12

# --add total score to list--#
score = []
# print(sheet_ranges.cell(row=row2, column=7).value)
row_i = 2
while True:  # row_num 2~11
    score.append(sheet_ranges.cell(row=row_i, column=7).value)
    row_i += 1
    if row_i is row_num:
        break

# score.reverse() //실수 적용이 안돼서 sorted 함수 사용.

list1 = sorted(score, key=float, reverse=True)
print(list1)
print(len(list1))
# --update grade-- #
row_i = 2

a = int(len(list1) * 0.3)
a_plus = int(a * 0.5)
a_zero = a - a_plus

b = int((len(list1) * 0.7) - a)
b_plus = int(b * 0.5)
b_zero = b - b_plus

c = int((len(list1)) - a - b)
c_plus = int(c * 0.5)
c_zero = c - c_plus

print(a_plus, a_zero, b_plus, b_zero, c_plus, c_zero)
while True:  # row_num 2~11
    v = list1.index(sheet_ranges.cell(row=row_i, column=7).value) + 1

    if v <= a_plus:
        g = 'A+'
    elif v <= a_plus + a_zero:
        g = 'A0'
    elif v <= a + b_plus:
        g = 'B+'
    elif v <= a + b_plus + b_zero:
        g = 'B0'
    elif v <= a + b + c_plus:
        g = 'C+'
    else:
        g = 'C0'

    sheet_ranges.cell(row=row_i, column=8, value=g)

    if sheet_ranges.cell(row=row_i, column=7).value < 40:  # 40점 미만이면 무조건 F
        sheet_ranges.cell(row=row_i, column=8, value='F')

    row_i += 1
    if row_i is row_num:
        break

# print(row_num)  # row num = 12

wb.save(filename="student.xlsx")  # file_save_update
wb.save(filename="output.xlsx")  # file_save for output
